from flask import Blueprint, render_template, redirect, request, flash, url_for, get_flashed_messages, session, send_file, make_response
from .modules.models import Transaction, TransactionNature, PaymentMode, TransactionCategory, AccountCategory
from app import db
from .modules.models import *
from .modules.db_queries import *
from .modules.activity_logger import log_activity, create_notification
from .auth import login_required
from sqlalchemy.orm import joinedload, validates
from sqlalchemy.exc import SQLAlchemyError
from collections import defaultdict
from sqlalchemy import func
import matplotlib.pyplot as plt
from io import BytesIO
import base64
import plotly.graph_objects as go
import plotly.express as px
from datetime import datetime, date
from werkzeug.security import generate_password_hash, check_password_hash
from .modules.transaction_utils import *
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from xhtml2pdf import pisa
from jinja2 import Template
import tempfile
import os

finance_manager_bp = Blueprint('finance_manager', __name__, url_prefix='/finmng/<int:user_id>')

@finance_manager_bp.before_request
def check_finance_manager_access():
    """Check if user is logged in before accessing any finance manager route."""
    if 'user_id' not in session:
        flash('Please log in to access this page.', 'error')
        return redirect(url_for('home.login'))
    
    # Check if user is accessing their own routes or has appropriate role
    if session.get('role') not in ['Finance Manager', 'Admin']:
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('home.index'))

@finance_manager_bp.route('/')
def finance_manager(user_id):
    events = get_events(user_id)
    user = User.query.get(user_id)
    return render_template('finance_manager/financemanager_dashboard.html', events=events, user=user, user_id=user_id)

@finance_manager_bp.route('/view_events')
def view_events(user_id):
    events = get_events(user_id)
    return render_template('finance_manager/view_events.html', events=events, user_id=user_id)    

@finance_manager_bp.route('/event_details/<int:event_id>')
def event_details(user_id, event_id):
    try:
        # Fetch transaction IDs for the given event
        transaction_ids = get_transaction_ids(event_id)
        print(transaction_ids)

        # Ensure transaction_ids is a list
        if not transaction_ids:
            transaction_ids = []

        # Calculate revenue and expenses
        revenue = get_revenue_total(event_id)
        expenses = get_expense_total(event_id)

        # Fetch category and mode totals with names
        category_totals = {}
        for category_id in get_all_transaction_category_ids():
            category_name = get_category_name(category_id)  # Fetch the name for the category ID
            category_totals[category_name] = get_category_total(event_id, category_id)

        mode_totals = {}
        for mode_id in get_all_payment_mode_ids():
            mode_name = get_mode_name(mode_id)  # Fetch the name for the mode ID
            mode_totals[mode_name] = get_mode_total(event_id, mode_id)

        # Fetch event details for the selected event (name, date, etc.)
        event_details = get_event_details(event_id)
        event = Event.query.get(event_id)

        # Render the template with calculated data
        return render_template(
            'finance_manager/event_details.html',
            event_id=event_id,
            user_id=user_id,
            revenue=revenue,
            expenses=expenses,
            category_totals=category_totals,
            mode_totals=mode_totals,
            event_details=event_details,  # Pass event details to the template
            event=event  # Pass event object for action buttons
        )
    except Exception as e:
        print(f"Error fetching event details: {e}")
        return redirect(url_for('finance_manager.view_events', user_id=user_id))


# Fetch Events Helper Function
def get_events(user_id):
    try:
        # Fetch events for the finance manager
        events = (
            db.session.query(Event)
            .filter(Event.Finance_Manager == user_id)
            .options(
                joinedload(Event.department),
                joinedload(Event.event_type),
                joinedload(Event.finance_manager)
            )
            .all()
        )

        # Group events by their status
        grouped_events = defaultdict(list)
        for event in events:
            status = determine_status(event.Date)
            grouped_events[status].append({
                "id": event.Event_ID,
                "event_name": event.Name,
                "status": status,
                "type": event.event_type.Event_Type_Name,
                "department": event.department.Name,
                "date": event.Date
            })

        return dict(grouped_events)

    except SQLAlchemyError as e:
        print("Error fetching events:", e)
        return {}

# Helper Function: Determine Event Status
def determine_status(event_date):
    today = date.today()
    if event_date > today:
        return "Upcoming"
    elif event_date == today:
        return "Ongoing"
    else:
        return "Completed"

def get_event_details(event_id):
    """Fetch details of an event based on Event ID."""
    event = Event.query.filter_by(Event_ID=event_id).first()
    if event:
        # Fetch related department and event type names from their respective models
        department_name = event.department.Name if event.department else "Unknown Department"
        event_type_name = event.event_type.Event_Type_Name if event.event_type else "Unknown Event Type"
        
        # Fetch finance and event managers' names
        finance_manager_name = event.finance_manager.Username if event.finance_manager else "No Finance Manager"
        event_manager_name = event.event_manager.Username if event.event_manager else "No Event Manager"

        return {
            "id": event_id,
            "name": event.Name,                     # Event Name
            "date": event.Date,                     # Event Date
            "department": department_name,          # Department Name
            "event_type": event_type_name,          # Event Type Name
            "days": event.Days,                     # Event Days
            "finance_manager": finance_manager_name,  # Finance Manager Name
            "event_manager": event_manager_name     # Event Manager Name
        }
    return {}


@finance_manager_bp.route('/event_visualization/<int:event_id>')
def event_visualization(user_id, event_id):
    try:
        event = Event.query.get(event_id)
        if not event:
            flash("Event not found.", "error")
            return redirect(url_for('finance_manager.view_events', user_id=user_id))
        
        event_name = event.Name
        
        # Fetch revenue and expense data
        revenue = get_revenue_total(event_id)
        expense = get_expense_total(event_id)
        profit_loss = revenue - expense

        # Fetch budget information
        budget = Budget.query.filter_by(Event_ID=event_id).first()
        budget_amount = float(budget.Amount) if budget else 0
        budget_remaining = budget_amount - expense
        budget_utilization = (expense / budget_amount * 100) if budget_amount > 0 else 0

        # Fetch category totals
        categories = get_all_transaction_category_ids()
        category_totals = []
        for category_id in categories:
            category_total = get_category_total(event_id, category_id)
            if category_total > 0:
                category_name = get_category_name(category_id)
                category_totals.append({'name': category_name, 'total': category_total})

        # Fetch mode totals
        modes = get_all_payment_mode_ids()
        mode_totals = []
        for mode_id in modes:
            mode_total = get_mode_total(event_id, mode_id)
            if mode_total > 0:
                mode_name = get_mode_name(mode_id)
                mode_totals.append({'name': mode_name, 'total': mode_total})

        # Fetch revenue and expense breakdown by nature
        revenue_transactions = Transaction.query.filter_by(Event_ID=event_id).join(TransactionNature).filter(
            TransactionNature.Nature_Name == 'Credit'
        ).all()
        
        expense_transactions = Transaction.query.filter_by(Event_ID=event_id).join(TransactionNature).filter(
            TransactionNature.Nature_Name == 'Debit'
        ).all()

        # Revenue breakdown by category
        revenue_by_category = {}
        for txn in revenue_transactions:
            cat_name = txn.transaction_category.Category_Name if txn.transaction_category else 'Uncategorized'
            revenue_by_category[cat_name] = revenue_by_category.get(cat_name, 0) + float(txn.Amount)

        # Expense breakdown by category
        expense_by_category = {}
        for txn in expense_transactions:
            cat_name = txn.transaction_category.Category_Name if txn.transaction_category else 'Uncategorized'
            expense_by_category[cat_name] = expense_by_category.get(cat_name, 0) + float(txn.Amount)

        # ===== CHART 1: Revenue vs Expense Bar Chart =====
        revenue_expense_fig = go.Figure()
        revenue_expense_fig.add_trace(go.Bar(
            x=['Revenue', 'Expenses', 'Net Profit/Loss'],
            y=[revenue, expense, profit_loss],
            marker_color=['#10b981', '#f59e0b', '#667eea' if profit_loss >= 0 else '#ef4444'],
            text=[f'₹{revenue:,.2f}', f'₹{expense:,.2f}', f'₹{profit_loss:,.2f}'],
            textposition='outside',
            textfont=dict(size=14, family='Arial', color='#2d3748'),
            hovertemplate='<b>%{x}</b><br>Amount: ₹%{y:,.2f}<extra></extra>'
        ))
        revenue_expense_fig.update_layout(
            title=dict(
                text='Financial Overview',
                font=dict(size=20, family='Arial', color='#2d3748', weight='bold')
            ),
            xaxis=dict(
                title='Category',
                titlefont=dict(size=16, family='Arial', color='#4a5568'),
                tickfont=dict(size=14, family='Arial', color='#4a5568')
            ),
            yaxis=dict(
                title='Amount (₹)',
                titlefont=dict(size=16, family='Arial', color='#4a5568'),
                tickfont=dict(size=14, family='Arial', color='#4a5568'),
                gridcolor='#e2e8f0'
            ),
            template='plotly_white',
            height=450,
            showlegend=False,
            plot_bgcolor='#f8fafc',
            paper_bgcolor='white',
            margin=dict(l=60, r=20, t=80, b=60),
            autosize=True
        )
        revenue_expense_fig_json = revenue_expense_fig.to_json()

        # ===== CHART 2: Budget Tracker Gauge Chart =====
        budget_fig = None
        if budget_amount > 0:
            budget_fig = go.Figure(go.Indicator(
                mode="gauge+number+delta",
                value=expense,
                domain={'x': [0, 1], 'y': [0, 1]},
                title=dict(
                    text=f"Budget Utilization<br><span style='font-size:0.7em'>Allocated: ₹{budget_amount:,.2f}</span>",
                    font=dict(size=18, family='Arial', color='#2d3748')
                ),
                delta={'reference': budget_amount, 'increasing': {'color': "#ef4444"}, 'decreasing': {'color': "#10b981"}},
                number=dict(prefix="₹", font=dict(size=28, family='Arial')),
                gauge={
                    'axis': {'range': [None, budget_amount * 1.2], 'tickfont': dict(size=12)},
                    'bar': {'color': "#667eea", 'thickness': 0.8},
                    'steps': [
                        {'range': [0, budget_amount * 0.7], 'color': "#d1fae5"},
                        {'range': [budget_amount * 0.7, budget_amount], 'color': "#fed7aa"},
                        {'range': [budget_amount, budget_amount * 1.2], 'color': "#fecaca"}
                    ],
                    'threshold': {
                        'line': {'color': "red", 'width': 4},
                        'thickness': 0.75,
                        'value': budget_amount
                    }
                }
            ))
            budget_fig.update_layout(
                height=350,
                template='plotly_white',
                paper_bgcolor='white',
                margin=dict(l=10, r=10, t=80, b=10),
                autosize=True
            )
            budget_fig = budget_fig.to_json()

        # ===== CHART 3: Budget vs Actual Spending =====
        budget_comparison_fig = None
        if budget_amount > 0:
            budget_comparison_fig = go.Figure()
            budget_comparison_fig.add_trace(go.Bar(
                name='Allocated',
                x=['Budget'],
                y=[budget_amount],
                marker_color='#667eea',
                text=[f'₹{budget_amount:,.2f}'],
                textposition='outside',
                textfont=dict(size=14, family='Arial'),
                hovertemplate='<b>Allocated</b><br>₹%{y:,.2f}<extra></extra>'
            ))
            budget_comparison_fig.add_trace(go.Bar(
                name='Spent',
                x=['Budget'],
                y=[expense],
                marker_color='#f59e0b',
                text=[f'₹{expense:,.2f}'],
                textposition='outside',
                textfont=dict(size=14, family='Arial'),
                hovertemplate='<b>Spent</b><br>₹%{y:,.2f}<extra></extra>'
            ))
            budget_comparison_fig.add_trace(go.Bar(
                name='Remaining',
                x=['Budget'],
                y=[budget_remaining if budget_remaining > 0 else 0],
                marker_color='#10b981',
                text=[f'₹{budget_remaining:,.2f}'],
                textposition='outside',
                textfont=dict(size=14, family='Arial'),
                hovertemplate='<b>Remaining</b><br>₹%{y:,.2f}<extra></extra>'
            ))
            budget_comparison_fig.update_layout(
                title=dict(
                    text='Budget Allocation vs Spending',
                    font=dict(size=20, family='Arial', color='#2d3748', weight='bold')
                ),
                xaxis=dict(
                    title='',
                    tickfont=dict(size=14, family='Arial')
                ),
                yaxis=dict(
                    title='Amount (₹)',
                    titlefont=dict(size=16, family='Arial', color='#4a5568'),
                    tickfont=dict(size=14, family='Arial', color='#4a5568'),
                    gridcolor='#e2e8f0'
                ),
                barmode='group',
                template='plotly_white',
                height=450,
                plot_bgcolor='#f8fafc',
                paper_bgcolor='white',
                legend=dict(
                    font=dict(size=14, family='Arial'),
                    orientation='h',
                    yanchor='bottom',
                    y=1.02,
                    xanchor='right',
                    x=1
                ),
                margin=dict(l=60, r=20, t=80, b=60),
                autosize=True
            )
            budget_comparison_fig = budget_comparison_fig.to_json()

        # ===== CHART 4: Profit/Loss Indicator =====
        profit_loss_fig = go.Figure(go.Indicator(
            mode="number+delta",
            value=profit_loss,
            title=dict(
                text="Net Profit/Loss",
                font=dict(size=18, family='Arial', color='#2d3748')
            ),
            delta={'reference': 0, 'position': "top"},
            domain={'x': [0, 1], 'y': [0.2, 0.8]},
            number=dict(
                prefix="₹",
                valueformat=",.2f",
                font=dict(size=36, family='Arial', color='#667eea' if profit_loss >= 0 else '#ef4444')
            )
        ))
        profit_loss_fig.update_layout(
            height=350,
            template='plotly_white',
            paper_bgcolor='white',
            margin=dict(l=10, r=10, t=60, b=10),
            autosize=True
        )
        profit_loss_fig_json = profit_loss_fig.to_json()

        # ===== CHART 5: Category Distribution Pie Chart =====
        category_fig = None
        if category_totals:
            category_fig = px.pie(
                names=[cat['name'] for cat in category_totals],
                values=[cat['total'] for cat in category_totals],
                title='Transaction Distribution by Category',
                color_discrete_sequence=px.colors.sequential.Purples_r
            )
            category_fig.update_traces(
                textposition='inside',
                textinfo='percent+label',
                textfont=dict(size=14, family='Arial', color='white'),
                hovertemplate='<b>%{label}</b><br>Amount: ₹%{value:,.2f}<br>Percentage: %{percent}<extra></extra>',
                marker=dict(line=dict(color='white', width=2))
            )
            category_fig.update_layout(
                height=450,
                template='plotly_white',
                title=dict(
                    font=dict(size=20, family='Arial', color='#2d3748', weight='bold')
                ),
                legend=dict(
                    font=dict(size=13, family='Arial'),
                    orientation='v',
                    yanchor='middle',
                    y=0.5
                ),
                paper_bgcolor='white',
                margin=dict(l=20, r=20, t=80, b=40),
                autosize=True
            )
            category_fig = category_fig.to_json()

        # ===== CHART 6: Payment Mode Distribution =====
        mode_fig = None
        if mode_totals:
            mode_fig = px.pie(
                names=[mode['name'] for mode in mode_totals],
                values=[mode['total'] for mode in mode_totals],
                title='Payment Mode Distribution',
                color_discrete_sequence=px.colors.sequential.Blues_r
            )
            mode_fig.update_traces(
                textposition='inside',
                textinfo='percent+label',
                textfont=dict(size=14, family='Arial', color='white'),
                hovertemplate='<b>%{label}</b><br>Amount: ₹%{value:,.2f}<br>Percentage: %{percent}<extra></extra>',
                marker=dict(line=dict(color='white', width=2))
            )
            mode_fig.update_layout(
                height=450,
                template='plotly_white',
                title=dict(
                    font=dict(size=20, family='Arial', color='#2d3748', weight='bold')
                ),
                legend=dict(
                    font=dict(size=13, family='Arial'),
                    orientation='v',
                    yanchor='middle',
                    y=0.5
                ),
                paper_bgcolor='white',
                margin=dict(l=20, r=20, t=80, b=40),
                autosize=True
            )
            mode_fig = mode_fig.to_json()

        # ===== CHART 7: Revenue Sources Breakdown =====
        revenue_breakdown_fig = None
        if revenue_by_category:
            categories_list = list(revenue_by_category.keys())
            values_list = list(revenue_by_category.values())
            
            revenue_breakdown_fig = go.Figure()
            revenue_breakdown_fig.add_trace(go.Bar(
                x=categories_list,
                y=values_list,
                marker_color='#10b981',
                text=[f'₹{v:,.2f}' for v in values_list],
                textposition='outside',
                textfont=dict(size=13, family='Arial'),
                hovertemplate='<b>%{x}</b><br>Revenue: ₹%{y:,.2f}<extra></extra>'
            ))
            revenue_breakdown_fig.update_layout(
                title=dict(
                    text='Revenue Sources by Category',
                    font=dict(size=20, family='Arial', color='#2d3748', weight='bold')
                ),
                xaxis=dict(
                    title='Category',
                    titlefont=dict(size=16, family='Arial', color='#4a5568'),
                    tickfont=dict(size=13, family='Arial', color='#4a5568'),
                    tickangle=-45
                ),
                yaxis=dict(
                    title='Amount (₹)',
                    titlefont=dict(size=16, family='Arial', color='#4a5568'),
                    tickfont=dict(size=14, family='Arial', color='#4a5568'),
                    gridcolor='#e2e8f0'
                ),
                template='plotly_white',
                height=450,
                showlegend=False,
                plot_bgcolor='#f8fafc',
                paper_bgcolor='white',
                margin=dict(l=60, r=20, t=80, b=100),
                autosize=True
            )
            revenue_breakdown_fig = revenue_breakdown_fig.to_json()

        # ===== CHART 8: Expense Breakdown =====
        expense_breakdown_fig = None
        if expense_by_category:
            categories_list = list(expense_by_category.keys())
            values_list = list(expense_by_category.values())
            
            expense_breakdown_fig = go.Figure()
            expense_breakdown_fig.add_trace(go.Bar(
                x=categories_list,
                y=values_list,
                marker_color='#f59e0b',
                text=[f'₹{v:,.2f}' for v in values_list],
                textposition='outside',
                textfont=dict(size=13, family='Arial'),
                hovertemplate='<b>%{x}</b><br>Expense: ₹%{y:,.2f}<extra></extra>'
            ))
            expense_breakdown_fig.update_layout(
                title=dict(
                    text='Expense Breakdown by Category',
                    font=dict(size=20, family='Arial', color='#2d3748', weight='bold')
                ),
                xaxis=dict(
                    title='Category',
                    titlefont=dict(size=16, family='Arial', color='#4a5568'),
                    tickfont=dict(size=13, family='Arial', color='#4a5568'),
                    tickangle=-45
                ),
                yaxis=dict(
                    title='Amount (₹)',
                    titlefont=dict(size=16, family='Arial', color='#4a5568'),
                    tickfont=dict(size=14, family='Arial', color='#4a5568'),
                    gridcolor='#e2e8f0'
                ),
                template='plotly_white',
                height=450,
                showlegend=False,
                plot_bgcolor='#f8fafc',
                paper_bgcolor='white',
                margin=dict(l=60, r=20, t=80, b=100),
                autosize=True
            )
            expense_breakdown_fig = expense_breakdown_fig.to_json()

        # Render template
        return render_template(
            'finance_manager/event_visualization.html',
            revenue_expense_fig=revenue_expense_fig_json,
            category_fig=category_fig,
            mode_fig=mode_fig,
            budget_fig=budget_fig,
            budget_comparison_fig=budget_comparison_fig,
            profit_loss_fig=profit_loss_fig_json,
            revenue_breakdown_fig=revenue_breakdown_fig,
            expense_breakdown_fig=expense_breakdown_fig,
            user_id=user_id,
            event_id=event_id,
            event_name=event_name,
            revenue=revenue,
            expense=expense,
            profit_loss=profit_loss,
            budget_amount=budget_amount,
            budget_remaining=budget_remaining,
            budget_utilization=budget_utilization
        )
    except Exception as e:
        print(f"Error in visualization: {e}")
        import traceback
        traceback.print_exc()
        flash(f"Error generating visualization: {str(e)}", "error")
        return redirect(url_for('finance_manager.event_details', event_id=event_id, user_id=user_id))
    
def get_event_name(event_id):
    event = db.session.query(Event).filter_by(Event_ID=event_id).first()
    return event.Name if event else "Unknown Event"

@finance_manager_bp.route('/download_excel/<int:event_id>')
def download_excel(user_id, event_id):
    try:
        event = Event.query.get(event_id)
        if not event:
            flash("Event not found.", "error")
            return redirect(url_for('finance_manager.event_details', user_id=user_id, event_id=event_id))
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Financial Summary"
        
        # Header styling
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Event details
        ws_summary['A1'] = "Event Financial Report"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary['A2'] = f"Event: {event.Name}"
        ws_summary['A3'] = f"Date: {event.Date.strftime('%B %d, %Y') if event.Date else 'N/A'}"
        ws_summary['A4'] = f"Department: {event.department.Name if event.department else 'N/A'}"
        ws_summary['A5'] = f"Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p')}"
        
        # Financial summary
        revenue = get_revenue_total(event_id)
        expense = get_expense_total(event_id)
        profit_loss = revenue - expense
        
        budget = Budget.query.filter_by(Event_ID=event_id).first()
        budget_amount = float(budget.Amount) if budget else 0
        
        ws_summary['A7'] = "Financial Metrics"
        ws_summary['A7'].font = header_font
        ws_summary['A7'].fill = header_fill
        ws_summary['B7'].fill = header_fill
        
        ws_summary['A8'] = "Total Revenue"
        ws_summary['B8'] = revenue
        ws_summary['B8'].number_format = '₹#,##0.00'
        
        ws_summary['A9'] = "Total Expenses"
        ws_summary['B9'] = expense
        ws_summary['B9'].number_format = '₹#,##0.00'
        
        ws_summary['A10'] = "Net Profit/Loss"
        ws_summary['B10'] = profit_loss
        ws_summary['B10'].number_format = '₹#,##0.00'
        
        if budget_amount > 0:
            ws_summary['A11'] = "Allocated Budget"
            ws_summary['B11'] = budget_amount
            ws_summary['B11'].number_format = '₹#,##0.00'
            
            ws_summary['A12'] = "Budget Remaining"
            ws_summary['B12'] = budget_amount - expense
            ws_summary['B12'].number_format = '₹#,##0.00'
        
        # Category breakdown
        categories = get_all_transaction_category_ids()
        row = 14
        ws_summary[f'A{row}'] = "Category Breakdown"
        ws_summary[f'A{row}'].font = header_font
        ws_summary[f'A{row}'].fill = header_fill
        ws_summary[f'B{row}'].fill = header_fill
        
        row += 1
        for category_id in categories:
            total = get_category_total(event_id, category_id)
            if total > 0:
                category_name = get_category_name(category_id)
                ws_summary[f'A{row}'] = category_name
                ws_summary[f'B{row}'] = total
                ws_summary[f'B{row}'].number_format = '₹#,##0.00'
                row += 1
        
        # Ledger Sheet
        ws_ledger = wb.create_sheet("Transaction Ledger")
        headers = ["Date", "Description", "Category", "Nature", "Payment Mode", "Amount"]
        
        for col, header in enumerate(headers, 1):
            cell = ws_ledger.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Fetch all transactions
        transactions = Transaction.query.filter_by(Event_ID=event_id).order_by(Transaction.Date).all()
        
        for row, txn in enumerate(transactions, 2):
            # Calculate total amount from transaction items
            total_amount = sum(item.Amount for item in txn.items) if txn.items else 0
            # Get first description or use party name
            description = txn.items[0].Description if txn.items and txn.items[0].Description else (txn.Party_Name or '')
            
            ws_ledger.cell(row=row, column=1, value=txn.Date.strftime('%Y-%m-%d') if txn.Date else 'N/A')
            ws_ledger.cell(row=row, column=2, value=description)
            ws_ledger.cell(row=row, column=3, value=txn.transaction_category.Category_Name if txn.transaction_category else 'N/A')
            ws_ledger.cell(row=row, column=4, value=txn.transaction_nature.Nature_Name if txn.transaction_nature else 'N/A')
            ws_ledger.cell(row=row, column=5, value=txn.payment_mode.Mode_Name if txn.payment_mode else 'N/A')
            
            amount_cell = ws_ledger.cell(row=row, column=6, value=float(total_amount))
            amount_cell.number_format = '₹#,##0.00'
            
            # Apply borders
            for col in range(1, 7):
                ws_ledger.cell(row=row, column=col).border = border
        
        # Auto-adjust column widths
        for ws in [ws_summary, ws_ledger]:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{event.Name.replace(' ', '_')}_Financial_Report.xlsx"
        
        # Log activity
        log_activity(
            user_id=user_id,
            action='generated',
            entity_type='Report',
            entity_id=event_id,
            description=f'Finance Manager generated Excel report for event "{event.Name}"'
        )
        
        # Notify Event Manager
        if event.Event_Manager:
            create_notification(
                user_id=event.Event_Manager,
                title='Financial Report Generated',
                message=f'Finance Manager has generated an Excel report for event "{event.Name}"',
                notification_type='info',
                event_id=event_id
            )
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        print(f"Error generating Excel: {e}")
        import traceback
        traceback.print_exc()
        flash("Error generating Excel report.", "error")
        return redirect(url_for('finance_manager.event_details', user_id=user_id, event_id=event_id))

@finance_manager_bp.route('/download_pdf/<int:event_id>')
def download_pdf(user_id, event_id):
    try:
        event = Event.query.get(event_id)
        if not event:
            flash("Event not found.", "error")
            return redirect(url_for('finance_manager.event_details', user_id=user_id, event_id=event_id))
        
        # Get financial data
        revenue = get_revenue_total(event_id)
        expense = get_expense_total(event_id)
        profit_loss = revenue - expense
        
        budget = Budget.query.filter_by(Event_ID=event_id).first()
        budget_amount = float(budget.Amount) if budget else 0
        
        # Category breakdown
        categories = get_all_transaction_category_ids()
        category_data = []
        for category_id in categories:
            total = get_category_total(event_id, category_id)
            if total > 0:
                category_data.append({
                    'name': get_category_name(category_id),
                    'total': total
                })
        
        # Transactions with calculated amounts
        transactions = Transaction.query.filter_by(Event_ID=event_id).order_by(Transaction.Date).all()
        
        # Prepare transaction data with amounts
        transaction_data = []
        for txn in transactions:
            total_amount = sum(item.Amount for item in txn.items) if txn.items else 0
            description = txn.items[0].Description if txn.items and txn.items[0].Description else (txn.Party_Name or '-')
            transaction_data.append({
                'date': txn.Date,
                'description': description,
                'category': txn.transaction_category.Category_Name if txn.transaction_category else 'N/A',
                'nature': txn.transaction_nature.Nature_Name if txn.transaction_nature else 'N/A',
                'amount': total_amount
            })
        
        # HTML template for PDF
        html_template = '''
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                @page { size: A4; margin: 1cm; }
                body { font-family: Arial, sans-serif; color: #333; }
                .header { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 20px; text-align: center; margin-bottom: 20px; }
                .header h1 { margin: 0; font-size: 24px; }
                .header p { margin: 5px 0 0 0; font-size: 12px; }
                .section { margin-bottom: 20px; }
                .section h2 { color: #667eea; border-bottom: 2px solid #667eea; padding-bottom: 5px; }
                table { width: 100%; border-collapse: collapse; margin-top: 10px; }
                th { background: #667eea; color: white; padding: 10px; text-align: left; }
                td { padding: 8px; border-bottom: 1px solid #ddd; }
                tr:nth-child(even) { background: #f8fafc; }
                .metric { display: inline-block; width: 48%; margin: 5px 0; }
                .metric-label { font-weight: bold; color: #667eea; }
                .metric-value { font-size: 18px; color: #2d3748; }
                .footer { margin-top: 30px; text-align: center; font-size: 10px; color: #999; }
            </style>
        </head>
        <body>
            <div class="header">
                <h1>Financial Report</h1>
                <p>{{ event.Name }}</p>
                <p>Generated on {{ now }}</p>
            </div>
            
            <div class="section">
                <h2>Event Information</h2>
                <p><strong>Event Name:</strong> {{ event.Name }}</p>
                <p><strong>Date:</strong> {{ event.Date.strftime('%B %d, %Y') if event.Date else 'N/A' }}</p>
                <p><strong>Department:</strong> {{ event.department.Name if event.department else 'N/A' }}</p>
                <p><strong>Event Type:</strong> {{ event.event_type.Event_Type_Name if event.event_type else 'N/A' }}</p>
            </div>
            
            <div class="section">
                <h2>Financial Summary</h2>
                <div class="metric">
                    <span class="metric-label">Total Revenue:</span>
                    <span class="metric-value">₹{{ "{:,.2f}".format(revenue) }}</span>
                </div>
                <div class="metric">
                    <span class="metric-label">Total Expenses:</span>
                    <span class="metric-value">₹{{ "{:,.2f}".format(expense) }}</span>
                </div>
                <div class="metric">
                    <span class="metric-label">Net Profit/Loss:</span>
                    <span class="metric-value" style="color: {% if profit_loss >= 0 %}#10b981{% else %}#ef4444{% endif %}">₹{{ "{:,.2f}".format(profit_loss) }}</span>
                </div>
                {% if budget_amount > 0 %}
                <div class="metric">
                    <span class="metric-label">Allocated Budget:</span>
                    <span class="metric-value">₹{{ "{:,.2f}".format(budget_amount) }}</span>
                </div>
                {% endif %}
            </div>
            
            {% if category_data %}
            <div class="section">
                <h2>Category Breakdown</h2>
                <table>
                    <thead>
                        <tr>
                            <th>Category</th>
                            <th style="text-align: right;">Amount</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for cat in category_data %}
                        <tr>
                            <td>{{ cat.name }}</td>
                            <td style="text-align: right;">₹{{ "{:,.2f}".format(cat.total) }}</td>
                        </tr>
                        {% endfor %}
                    </tbody>
                </table>
            </div>
            {% endif %}
            
            <div class="section">
                <h2>Transaction Ledger</h2>
                <table>
                    <thead>
                        <tr>
                            <th>Date</th>
                            <th>Description</th>
                            <th>Category</th>
                            <th>Nature</th>
                            <th style="text-align: right;">Amount</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for txn in transaction_data %}
                        <tr>
                            <td>{{ txn.date.strftime('%Y-%m-%d') if txn.date else 'N/A' }}</td>
                            <td>{{ txn.description }}</td>
                            <td>{{ txn.category }}</td>
                            <td>{{ txn.nature }}</td>
                            <td style="text-align: right;">₹{{ "{:,.2f}".format(txn.amount) }}</td>
                        </tr>
                        {% endfor %}
                    </tbody>
                </table>
            </div>
            
            <div class="footer">
                <p>This is a computer-generated report from FinSight - Event Financial Management System</p>
            </div>
        </body>
        </html>
        '''
        
        template = Template(html_template)
        html_content = template.render(
            event=event,
            revenue=revenue,
            expense=expense,
            profit_loss=profit_loss,
            budget_amount=budget_amount,
            category_data=category_data,
            transaction_data=transaction_data,
            now=datetime.now().strftime('%B %d, %Y %I:%M %p')
        )
        
        # Generate PDF using xhtml2pdf
        pdf_file = BytesIO()
        pisa_status = pisa.CreatePDF(
            html_content,
            dest=pdf_file
        )
        
        if pisa_status.err:
            raise Exception("Error generating PDF")
        
        pdf_file.seek(0)
        filename = f"{event.Name.replace(' ', '_')}_Financial_Report.pdf"
        
        # Log activity
        log_activity(
            user_id=user_id,
            action='generated',
            entity_type='Report',
            entity_id=event_id,
            description=f'Finance Manager generated PDF report for event "{event.Name}"'
        )
        
        # Notify Event Manager
        if event.Event_Manager:
            create_notification(
                user_id=event.Event_Manager,
                title='PDF Report Generated',
                message=f'Finance Manager has generated a PDF report for event "{event.Name}"',
                notification_type='info',
                event_id=event_id
            )
        
        return send_file(
            pdf_file,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        print(f"Error generating PDF: {e}")
        import traceback
        traceback.print_exc()
        flash("Error generating PDF report.", "error")
        return redirect(url_for('finance_manager.event_details', user_id=user_id, event_id=event_id))

@finance_manager_bp.route('/ledger/<int:event_id>')
def ledger_view(user_id, event_id):
    try:
        event = Event.query.get(event_id)
        if not event:
            flash("Event not found.", "error")
            return redirect(url_for('finance_manager.view_events', user_id=user_id))
        
        # Fetch all transactions with related data
        transactions = Transaction.query.filter_by(Event_ID=event_id).order_by(Transaction.Date.desc()).all()
        
        # Calculate running balance
        running_balance = 0
        ledger_entries = []
        
        for txn in reversed(transactions):
            # Calculate total amount from transaction items
            total_amount = sum(item.Amount for item in txn.items) if txn.items else 0
            
            if txn.transaction_nature and txn.transaction_nature.Nature_Name == 'Credit':
                running_balance += float(total_amount)
            else:
                running_balance -= float(total_amount)
            
            ledger_entries.append({
                'transaction': txn,
                'balance': running_balance,
                'amount': total_amount
            })
        
        ledger_entries.reverse()
        
        # Get totals
        revenue = get_revenue_total(event_id)
        expense = get_expense_total(event_id)
        
        return render_template(
            'finance_manager/ledger_view.html',
            event=event,
            ledger_entries=ledger_entries,
            revenue=revenue,
            expense=expense,
            user_id=user_id
        )
    
    except Exception as e:
        print(f"Error in ledger view: {e}")
        import traceback
        traceback.print_exc()
        flash("Error loading ledger.", "error")
        return redirect(url_for('finance_manager.event_details', user_id=user_id, event_id=event_id))


@finance_manager_bp.route('/profile')
def finance_manager_profile(user_id):
    """Display finance manager profile page"""
    try:
        user = User.query.get_or_404(user_id)
        
        # Get event count for this finance manager
        event_count = Event.query.filter_by(Finance_Manager=user_id).count()
        
        # Count sub-events under the events managed by this finance manager
        sub_event_count = db.session.query(SubEvent).join(Event).filter(
            Event.Finance_Manager == user_id
        ).count()
        
        return render_template('finance_manager/profile.html', 
                             user=user, 
                             user_id=user_id,
                             event_count=event_count,
                             sub_event_count=sub_event_count)
    except Exception as e:
        flash(f"Error loading profile: {str(e)}", "danger")
        return redirect(url_for('finance_manager.finance_manager', user_id=user_id))


@finance_manager_bp.route('/profile/edit', methods=['GET', 'POST'])
def edit_finance_manager_profile(user_id):
    """Edit finance manager profile"""
    try:
        user = User.query.get_or_404(user_id)
        
        if request.method == 'POST':
            username = request.form.get('username')
            email = request.form.get('email')
            current_password = request.form.get('current_password')
            new_password = request.form.get('new_password')
            confirm_password = request.form.get('confirm_password')
            
            old_username = user.Username
            old_email = user.Email
            
            user.Username = username
            user.Email = email
            
            if current_password and new_password:
                if not check_password_hash(user.Password, current_password):
                    flash("Current password is incorrect", "danger")
                    return redirect(url_for('finance_manager.edit_finance_manager_profile', user_id=user_id))
                
                if new_password != confirm_password:
                    flash("New passwords do not match", "danger")
                    return redirect(url_for('finance_manager.edit_finance_manager_profile', user_id=user_id))
                
                user.Password = generate_password_hash(new_password)
                log_activity(
                    user_id=user_id,
                    action='updated',
                    entity_type='User',
                    entity_id=user_id,
                    description=f"Finance Manager '{user.Username}' changed their password"
                )
            
            db.session.commit()
            
            changes = []
            if old_username != username:
                changes.append(f"username from '{old_username}' to '{username}'")
            if old_email != email:
                changes.append(f"email from '{old_email}' to '{email}'")
            
            if changes:
                log_activity(
                    user_id=user_id,
                    action='updated',
                    entity_type='User',
                    entity_id=user_id,
                    description=f"Finance Manager updated profile: {', '.join(changes)}"
                )
            
            flash("Profile updated successfully", "success")
            return redirect(url_for('finance_manager.finance_manager_profile', user_id=user_id))
        
        return render_template('finance_manager/edit_profile.html', user=user, user_id=user_id)
    
    except Exception as e:
        db.session.rollback()
        flash(f"Error updating profile: {str(e)}", "danger")
        return redirect(url_for('finance_manager.finance_manager_profile', user_id=user_id))


@finance_manager_bp.route('/notifications')
def notifications(user_id):
    """Display notifications for finance manager"""
    try:
        # Get all notifications for this user, ordered by most recent
        notifications_list = Notification.query.filter_by(User_ID=user_id).order_by(Notification.Created_At.desc()).all()
        unread_count = Notification.query.filter_by(User_ID=user_id, Is_Read=False).count()
        
        return render_template('finance_manager/notifications.html', 
                             notifications=notifications_list,
                             unread_count=unread_count,
                             user_id=user_id)
    except Exception as e:
        flash(f"Error loading notifications: {str(e)}", "danger")
        return redirect(url_for('finance_manager.finance_manager', user_id=user_id))


@finance_manager_bp.route('/notifications/<int:notification_id>/read', methods=['POST'])
def mark_notification_read(user_id, notification_id):
    """Mark a single notification as read"""
    try:
        from .modules.activity_logger import mark_notification_read
        mark_notification_read(notification_id)
        flash("Notification marked as read", "success")
    except Exception as e:
        flash(f"Error: {str(e)}", "danger")
    return redirect(url_for('finance_manager.notifications', user_id=user_id))


@finance_manager_bp.route('/notifications/mark-all-read', methods=['POST'])
def mark_all_notifications_read(user_id):
    """Mark all notifications as read"""
    try:
        from .modules.activity_logger import mark_all_notifications_read
        mark_all_notifications_read(user_id)
        flash("All notifications marked as read", "success")
    except Exception as e:
        flash(f"Error: {str(e)}", "danger")
    return redirect(url_for('finance_manager.notifications', user_id=user_id))
